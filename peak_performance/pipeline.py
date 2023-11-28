"""
PeakPerformance
Copyright (C) 2023 Forschungszentrum Jülich GmbH

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU Affero General Public License as published
by the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU Affero General Public License for more details.

You should have received a copy of the GNU Affero General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

import importlib
import os
import re
import shutil
import warnings
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Mapping, Sequence, Tuple, Union

import arviz as az
import numpy as np
import pandas
import pymc as pm
import scipy.integrate
import scipy.signal
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from peak_performance import models, plots


class ParsingError(Exception):
    """Base type of parsing exceptions."""


class InputError(Exception):
    """Base type of exceptions related to information given by the user."""


class UserInput:
    """Collect all information required from the user and format them in the correct manner."""

    def __init__(
        self,
        path: Union[str, os.PathLike],
        files: Sequence[str],
        raw_data_file_format: str,
        peak_model: Sequence[str],
        retention_time_estimate: Union[Sequence[float], Sequence[int]],
        peak_width_estimate: Union[float, int],
        pre_filtering: bool,
        minimum_sn: Union[float, int],
        timeseries: np.ndarray,
        acquisition: str,
        precursor: Union[float, int],
        product_mz_start: Union[float, int],
        product_mz_end: Union[float, int],
    ):
        """
        Parameters
        ----------
        path
            Path to the folder containing the results of the current run.
        files
            List of raw data file names in path.
        raw_data_file_format
            Data format (suffix) of the raw data, default is '.npy'.
        peak_model
            List specifying models for peak fitting in the same order as files.
            ("normal", "skew_normal", "double_normal", "double_skew_normal")
        retention_time_estimate
            In case you set pre_filtering to True, give a retention time estimate (float) for each signal in files.
            In case of a double peak, give two retention times (in chronological order) as a tuple containing two floats.
        peak_width_estimate
            Rough estimate of the average peak width in minutes expected for the LC-MS method with which the data was obtained.
        pre_filtering
            If True, potential peaks will be filtered based on retention time and signal to noise ratio before sampling.
        minimum_sn
            Minimum signal to noise ratio for a signal to be recognized as a peak during pre-filtering.
        timeseries
            NumPy Array containing time (at first position) and intensity (at second position) data as NumPy arrays.
        acquisition
            Name of a single acquisition.
        precursor
            Can be one of the following:
            Either the experiment number of the signal within the acquisition (each experiment = one mass trace)
            or the mass to charge ratio of the precursor ion selected in Q1.
        product_mz_start
            Start of the mass to charge ratio range of the product ion in the TOF.
        product_mz_end
            End of the mass to charge ratio range of the product ion in the TOF.
        """
        self.path = path
        self.files = list(files)
        self.raw_data_file_format = raw_data_file_format
        self.peak_model = peak_model
        self.retention_time_estimate = retention_time_estimate
        self.peak_width_estimate = peak_width_estimate
        self.pre_filtering = pre_filtering
        self.minimum_sn = minimum_sn
        self.timeseries = timeseries
        self.acquisition = acquisition
        self.precursor = precursor
        self.product_mz_start = product_mz_start
        self.product_mz_end = product_mz_end
        super().__init__()

    @property
    def timeseries(self):
        """
        Getting the value of the timeseries attribute.
        (NumPy Array containing time (at first position) and intensity (at second position) data as NumPy arrays.)
        """
        return self._timeseries

    @timeseries.setter
    def timeseries(self, data):
        """Setting the value of the timeseries attribute."""
        if data is None:
            raise InputError("The timeseries parameter is a None type.")
        self._timeseries = np.asarray(data)

    @property
    def acquisition(self):
        """Getting the value of the acquisition attribute (name of a single acquisition)."""
        return self._acquisition

    @acquisition.setter
    def acquisition(self, name):
        """Setting the value of the acquisition attribute."""
        if not isinstance(name, str):
            raise InputError(
                f"The acquisition parameter {name} is {type(name)} but needs to be a string."
            )
        if name is None:
            raise InputError("The acquisition parameter is a None type.")
        self._acquisition = name

    @property
    def precursor(self):
        """
        Getting the value of the precursor attribute which can be one of the following:
            Either the experiment number of the signal within the acquisition (each experiment = one mass trace)
            or the mass to charge ratio of the precursor ion selected in Q1.
        """
        return self._precursor

    @precursor.setter
    def precursor(self, mz):
        """Setting the value of the precursor attribute."""
        if not isinstance(mz, int) and not isinstance(mz, float):
            try:
                mz = float(mz)
            except ValueError as ex:
                raise InputError(
                    f"The precursor parameter {mz} is {type(mz)} but needs to be an int or a float."
                ) from ex
        if mz is None:
            raise InputError("The precursor parameter is a None type.")
        self._precursor = mz

    @property
    def product_mz_start(self):
        """Getting the value of the product_mz_start attribute."""
        return self._product_mz_start

    @product_mz_start.setter
    def product_mz_start(self, mz):
        """
        Setting the value of the product_mz_start attribute.
        (Start of the mass to charge ratio range of the product ion in the TOF.)
        """
        if not isinstance(mz, int) and not isinstance(mz, float):
            try:
                mz = float(mz)
            except ValueError as ex:
                raise InputError(
                    f"The product_mz parameter {mz} is {type(mz)} but needs to be an int or a float."
                ) from ex
        if mz is None:
            raise InputError("The product_mz_start parameter is a None type.")
        self._product_mz_start = mz

    @property
    def product_mz_end(self):
        """
        Getting the value of the product_mz_end attribute.
        (End of the mass to charge ratio range of the product ion in the TOF.)
        """
        return self._product_mz_end

    @product_mz_end.setter
    def product_mz_end(self, mz):
        """Setting the value of the product_mz_end attribute."""
        if not isinstance(mz, int) and not isinstance(mz, float):
            try:
                mz = float(mz)
            except ValueError as ex:
                raise InputError(
                    f"The product_mz parameter is {type(mz)} but needs to be an int or a float."
                ) from ex
        if mz is None:
            raise InputError("The product_mz_end parameter is a None type.")
        self._product_mz_end = mz

    @property
    def user_info(self):
        """Create a dictionary with the necessary user information based on the class attributes."""
        # first, some sanity checks
        if len(self.files) != len(self.peak_model):
            raise InputError(
                f"The length of 'files' ({len(self.files)}) and of 'peak_model' ({len(self.peak_model)}) are not identical."
            )
        if self.pre_filtering:
            # check length of lists
            if len(self.files) != len(self.peak_model) or len(self.peak_model) != len(
                self.retention_time_estimate
            ):
                raise InputError(
                    f"The length of 'files' ({len(self.files)}), 'peak_model' ({self.peak_model}), "
                    f"and retention_time_estimate ({len(self.retention_time_estimate)}) are not identical."
                )
        else:
            # if pre_filtering is False, then retention_time_estimate is not needed
            # but the dictionary still needs to be created without errors -> set it to np.nan
            if len(self.retention_time_estimate) == 1:
                self.retention_time_estimate = len(self.files) * [np.nan]
            elif not self.retention_time_estimate:
                self.retention_time_estimate = len(self.files) * [np.nan]
        if np.any(np.array(self.retention_time_estimate) < 0):
            raise InputError("Retention time estimates below 0 are not valid.")
        # actually create the dictionary
        user_info = dict(zip(self.files, zip(self.peak_model, self.retention_time_estimate)))
        user_info["peak_width_estimate"] = self.peak_width_estimate
        user_info["pre_filtering"] = self.pre_filtering
        user_info["minimum_sn"] = self.minimum_sn
        return user_info


def detect_raw_data(path: Union[str, os.PathLike], *, data_type: str = ".npy"):
    """
    Detect all .npy files with time and intensity data for peaks in a given directory.

    Parameters
    ----------
    path
        Path to the folder containing raw data.
    data_type
        Data format of the raw data files (e.g. '.npy').

    Returns
    -------
    files
        List with names of all files of the specified data type in path.
    """
    all_files = os.listdir(path)
    files = [file for file in all_files if data_type in file]
    if not files:
        raise FileNotFoundError(
            f"In the given directory '{path}', there are no '{data_type}' files."
        )
    return files


def parse_data(
    path: Union[str, os.PathLike], filename: str, raw_data_file_format: str
) -> Tuple[np.ndarray, str, float, float, float]:
    """
    Extract names of data files.

    Parameters
    ----------
    path
        Path to the raw data files.
    filename
        Name of a raw date file containing a NumPy array with a time series (time as first, intensity as second element of the array).
    raw_data_file_format
        Data format (suffix) of the raw data, default is '.npy'.

    Returns
    -------
    timeseries
        Updated NumPy array containing time and intensity data as NumPy arrays in first and second row, respectively.
        NaN values have been replaced with zeroes.
    acquisition
        Name of a single acquisition.
    precursor
        Can be one of the following:
        Either the experiment number of the signal within the acquisition (each experiment = one mass trace)
        or the mass to charge ratio of the precursor ion selected in Q1.
    product_mz_start
        Start of the mass to charge ratio range of the product ion in the TOF.
    product_mz_end
        End of the mass to charge ratio range of the product ion in the TOF.
    """
    # load time series
    timeseries = np.load(Path(path) / filename)
    # if NaN are in time or intensity, replace it with 0.0
    timeseries = np.nan_to_num(timeseries)
    # get information from the raw data file name
    splits = filename.split("_")
    if len(splits) != 4:
        raise InputError(
            f"""The standardized naming scheme was violated by file {filename}.
            \nThe name should be divided by underscores into the sections acquisition name, precursor, product_mz_start, and product_mz_end.
            """
        )
    try:
        pattern = r"(.*?)_(\d+\.?\d*)_(\d+\.?\d*)_(\d+\.?\d*).*"
        m = re.match(pattern, filename)
        if m is not None:
            acquisition, precursor, mz_start, mz_end = m.groups()
        precursor_converted = float(precursor)
        product_mz_start_converted = float(mz_start)
        product_mz_end_converted = float(mz_end)
    except ValueError as ex:
        raise InputError(
            f"The name of file {filename} does not follow the standardized naming convention."
        ) from ex

    return (
        timeseries,
        acquisition,
        precursor_converted,
        product_mz_start_converted,
        product_mz_end_converted,
    )


def parse_unique_identifiers(raw_data_files: Sequence[str]) -> List[str]:
    """
    Get a set of all mass traces based on the standardized raw data file names (excluding acquisitions).
    Used to automatically fill out the unique_identifiers column in the Template.xlsx' signals tab.

    Parameters
    ----------
    raw_data_files
        Names of all files of the specified data type in path_raw_data.

    Returns
    -------
    unique_identifiers
        List with all unique combinations of targeted molecules.
        (i.e. experiment number or precursor ion m/z ratio and product ion m/z ratio range)
    """
    # remove acquisition from file names
    identifiers = []
    for filename in raw_data_files:
        pattern = r"(.*?)_(\d+\.?\d*)_(\d+\.?\d*)_(\d+\.?\d*).*"
        m = re.match(pattern, filename)
        if m is not None:
            acquisition, precursor, mz_start, mz_end = m.groups()
        identifiers.append("_".join([precursor, mz_start, mz_end]))

    # select only unique identifiers
    unique_identifiers = list(set(identifiers))
    return unique_identifiers


def initiate(path: Union[str, os.PathLike], *, run_dir: str = ""):
    """
    Create a folder for the results. Also create a zip file inside that folder. Also create df_summary.

    Parameters
    ----------
    path
        Path to the directory containing the raw data.
    run_dir
        Name of the directory created to store the results of the current run (default: current date and time).

    Returns
    -------
    df_summary
        DataFrame for collecting the results (i.e. peak parameters) of every signal of a given pipeline.
    path
        Updated path variable pointing to the newly created folder for this batch.
    """
    # get current date and time
    if not run_dir:
        today = str(date.today())
        now = datetime.now().strftime("%H-%M-%S")
        timestamp = today + "_" + now
        run_dir = timestamp + "_run"
        # create a directory
    path = Path(path) / run_dir
    path.mkdir(exist_ok=True)
    # create DataFrame for data report
    df_summary = pandas.DataFrame(
        columns=[
            "mean",
            "sd",
            "hdi_3%",
            "hdi_97%",
            "mcse_mean",
            "mcse_sd",
            "ess_bulk",
            "ess_tail",
            "r_hat",
            "acquisition",
            "experiment_or_precursor_mz",
            "product_mz_start",
            "product_mz_end",
            "is_peak",
            "cause_for_rejection",
            "model_type",
            "subpeak",
        ]
    )
    return df_summary, path


def prefiltering(
    filename: str, ui: UserInput, noise_width_guess: float, df_summary: pandas.DataFrame
):
    """
    Optional method to skip signals where clearly no peak is present. Saves a lot of computation time.

    Parameters
    ----------
    filename
        Name of the raw data file.
    ui
        Instance of the UserInput class
    noise_width_guess
        Estimated width of the noise of a particular measurement.

    Returns
    -------
    found_peak
        True, if any peak candidate was found within the time frame; False, if not.
    df_summary
        DataFrame for collecting the results (i.e. peak parameters) of every signal of a given pipeline.
    """
    # pre-fit tests for peaks to save computation time (optional)
    t_ret = ui.user_info[filename][1]
    est_width = ui.peak_width_estimate
    # find all potential peaks with scipy
    peaks, _ = scipy.signal.find_peaks(ui.timeseries[1])
    peak_candidates = []
    # differentiate between single and double peaks
    for peak in peaks:
        # define conditions for passing the pre-filtering
        # check proximity of any peak candidate to the estimated retention time
        retention_time_condition = t_ret - est_width <= ui.timeseries[0][peak] <= t_ret + est_width
        # check signal to noise ratio
        signal_to_noise_condition = (
            ui.timeseries[1][peak] / (noise_width_guess + 0.1) > ui.minimum_sn
        )
        # check the neighbouring data points to prevent classification of a single elevated data point as a peak
        check_preceding_point = ui.timeseries[1][peak - 1] / (noise_width_guess + 0.1) > 2
        check_succeeding_point = ui.timeseries[1][peak + 1] / (noise_width_guess + 0.1) > 2
        if (
            retention_time_condition
            and signal_to_noise_condition
            and check_preceding_point
            and check_succeeding_point
        ):
            peak_candidates.append(peak)
    if not peak_candidates:
        df_summary = report_add_nan_to_summary(filename, ui, df_summary, "pre-filtering")
        return False, df_summary
    return True, df_summary


def sampling(pmodel, **sample_kwargs):
    """Performs sampling.

    Parameters
    ----------
    pmodel
        A PyMC model.
    **kwargs
        The keyword arguments are used in pm.sample().
    tune
        Number of tuning samples (default = 2000).
    draws
        Number of samples after tuning (default = 2000).

    Returns
    -------
    idata
        Inference data object.
    """
    sample_kwargs.setdefault("tune", 2000)
    sample_kwargs.setdefault("draws", 2000)
    # check if nutpie is available; if so, use it to enhance performance
    if importlib.util.find_spec("nutpie"):
        nuts_sampler = "nutpie"
    else:
        nuts_sampler = "pymc"
    with pmodel:
        idata = pm.sample_prior_predictive()
        idata.extend(pm.sample(nuts_sampler=nuts_sampler, **sample_kwargs))
    return idata


def postfiltering(filename: str, idata, ui: UserInput, df_summary: pandas.DataFrame):
    """
    Method to filter out false positive peaks after sampling based on the obtained uncertainties of several peak parameters.

    Parameters
    ----------
    filename
        Name of the raw data file.
    idata
        Inference data object resulting from sampling.
    ui
        Instance of the UserInput class.
    df_summary
        DataFrame for collecting the results (i.e. peak parameters) of every signal of a given pipeline.

    Returns
    -------
    acceptance
        True if the signal was accepted as a peak -> save data and continue with next signal.
        False if the signal was not accepted as a peak -> re-sampling with more tuning samples or discard signal.
    resample
        True: re-sample with more tuning samples, False: don't.
    discard
        True: discard sample.
    """
    # check whether convergence, i.e. r_hat <= 1.05, was not reached OR peak criteria were not met
    model = ui.user_info[filename][0]
    resample = False
    discard = False
    rejection_msg = ""
    az_summary: pandas.DataFrame = az.summary(idata)
    if model in ["normal", "skew_normal"]:
        # for single peak
        # Get data needed for rejection decisions
        max_rhat = max(az_summary.loc[:, "r_hat"])
        std = az_summary.loc["std", "mean"]
        area_sd = az_summary.loc["area", "sd"]
        area_mean = az_summary.loc["area", "mean"]
        height_sd = az_summary.loc["height", "sd"]
        height_mean = az_summary.loc["height", "mean"]

        # decide whether to discard signal or sample with more tune samples based on size of sigma parameter
        # of normal distribution (std) and on the relative sizes of standard deviations of area and height
        reject_reasons = []
        if max_rhat > 1.05:
            reject_reasons.append(f"maximum Rhat ({max_rhat:.3f}) was too high")
        if std <= ui.peak_width_estimate / 100:
            reject_reasons.append(f"standard deviation estimate ({std:.2f}) was too low")
        if area_sd > area_mean * 0.2:
            reject_reasons.append(f"area estimate ({area_mean} ± {area_sd}) was too uncertain")
        if height_sd > height_mean * 0.2:
            reject_reasons.append(
                f"height estimate ({height_mean} ± {height_sd}) was too uncertain"
            )

        if len(reject_reasons) == 1 and "Rhat" in reject_reasons[0]:
            # r_hat failed but rest of post-fit check passed
            # sample again with more tune samples to possibly reach convergence yet
            resample = True
            discard = False
        elif reject_reasons:
            rejection_msg = " and ".join(reject_reasons)
            df_summary = report_add_nan_to_summary(filename, ui, df_summary, rejection_msg)
            resample = False
            discard = True

    elif model in ["double_normal", "double_skew_normal"]:
        # for double peak
        max_rhat = max(az_summary.loc[:, "r_hat"])
        std = az_summary.loc["std[0]", "mean"]
        area_sd = az_summary.loc["area[0]", "sd"]
        area_mean = az_summary.loc["area[0]", "mean"]
        height_sd = az_summary.loc["height[0]", "sd"]
        height_mean = az_summary.loc["height[0]", "mean"]
        std2 = az_summary.loc["std[1]", "mean"]
        area_sd2 = az_summary.loc["area[1]", "sd"]
        area_mean2 = az_summary.loc["area[1]", "mean"]
        height_sd2 = az_summary.loc["height[1]", "sd"]
        height_mean2 = az_summary.loc["height[1]", "mean"]

        if max_rhat > 1.05:
            resample = True
            discard = False
            return resample, discard, df_summary
        # Booleans to differentiate which peak is or is not detected
        double_not_found_first = False
        double_not_found_second = False
        if std <= 1 / 100 or area_sd > area_mean * 0.2 or height_sd > height_mean * 0.2:
            # post-fit check failed
            # add NaN values to summary DataFrame
            double_not_found_first = True
        if std2 <= 1 / 100 or area_sd2 > area_mean2 * 0.2 or height_sd2 > height_mean2 * 0.2:
            # post-fit check failed
            # add NaN values to summary DataFrame
            double_not_found_second = True
        # if both peaks failed the peak criteria tests, then reject peaks
        if double_not_found_first and double_not_found_second:
            reject_reasons = []
            if std <= ui.peak_width_estimate / 100:
                reject_reasons.append(f"standard deviation estimate ({std:.2f}) was too low")
            if std2 <= ui.peak_width_estimate / 100:
                reject_reasons.append(f"standard deviation estimate ({std2:.2f}) was too low")
            if area_sd > area_mean * 0.2:
                reject_reasons.append(f"area estimate ({area_mean} ± {area_sd}) was too uncertain")
            if area_sd2 > area_mean2 * 0.2:
                reject_reasons.append(
                    f"area estimate ({area_mean2} ± {area_sd2}) was too uncertain"
                )
            if height_sd > height_mean * 0.2:
                reject_reasons.append(
                    f"height estimate ({height_mean} ± {height_sd}) was too uncertain"
                )
            if height_sd2 > height_mean2 * 0.2:
                reject_reasons.append(
                    f"height estimate ({height_mean2} ± {height_sd2}) was too uncertain"
                )

            if reject_reasons:
                rejection_msg = " and ".join(reject_reasons)

            df_summary = report_add_nan_to_summary(filename, ui, df_summary, rejection_msg)
            resample = False
            discard = True

    else:
        raise NotImplementedError(f"The model {model} is not implemented.")
    return resample, discard, df_summary


def posterior_predictive_sampling(pmodel, idata):
    """Performs posterior predictive sampling for signals recognized as peaks.

    Parameters
    ----------
    pmodel
        A PyMC model.
    idata
        Previously sampled inference data object.

    Returns
    -------
    idata
        Inference data object updated with the posterior predictive samples.
    """
    with pmodel:
        idata.extend(pm.sample_posterior_predictive(idata, var_names=["y"]))
    return idata


def report_save_idata(idata, ui: UserInput, filename: str):
    """
    Saves inference data object as a .nc file.

    Parameters
    ----------
    idata
        Inference data object resulting from sampling.
    ui
        Instance of the UserInput class.
    filename
        Name of a raw date file containing a NumPy array with a time series (time as first, intensity as second element of the array).
    """
    fp = Path(ui.path) / f"{filename}.nc"
    idata.to_netcdf(str(fp.absolute()))
    return


def report_add_data_to_summary(
    filename: str,
    idata,
    df_summary: pandas.DataFrame,
    ui: UserInput,
    is_peak: bool,
    rejection_cause: str = "",
):
    """
    Extracts the relevant information from idata, concatenates it to the summary DataFrame, and saves the DataFrame as an Excel file.
    Error handling prevents stop of the pipeline in case the saving doesn't work (e.g. because the file was opened by someone).

    Parameters
    ----------
    idata
        Inference data object resulting from sampling.
    df_summary
        DataFrame for collecting the results (i.e. peak parameters) of every signal of a given pipeline.
    ui
        Instance of the UserInput class.
    is_peak
        Boolean stating whether a signal was recognized as a peak (True) or not (False).
    rejection_cause
        Cause for rejecting a given signal.

    Returns
    -------
    df_summary
        Updated DataFrame for collecting the results (i.e. peak parameters) of every signal of a given pipeline.
    """
    az_summary: pandas.DataFrame = az.summary(idata)
    model = ui.user_info[filename][0]
    # split double peak into first and second peak (when extracting the data from az.summary(idata))
    if model in ["double_normal", "double_skew_normal"]:
        # first peak of double peak
        parameters = [
            "baseline_intercept",
            "baseline_slope",
            "mean[0]",
            "noise",
            "std[0]",
            "area[0]",
            "height[0]",
            "sn[0]",
        ]
        df = az_summary.loc[parameters, :]
        df = df.rename(
            index={
                "mean[0]": "mean",
                "std[0]": "std",
                "area[0]": "area",
                "height[0]": "height",
                "sn[0]": "sn",
            }
        )
        df["acquisition"] = len(parameters) * [f"{ui.acquisition}"]
        df["experiment_or_precursor_mz"] = len(parameters) * [ui.precursor]
        df["product_mz_start"] = len(parameters) * [ui.product_mz_start]
        df["product_mz_end"] = len(parameters) * [ui.product_mz_end]
        df["is_peak"] = is_peak
        df["cause_for_rejection"] = rejection_cause
        df["model_type"] = len(parameters) * [model]
        df["subpeak"] = len(parameters) * ["1st"]

        # second peak of double peak
        parameters = [
            "baseline_intercept",
            "baseline_slope",
            "mean[1]",
            "noise",
            "std[1]",
            "area[1]",
            "height[1]",
            "sn[1]",
        ]
        df2 = az_summary.loc[parameters, :]
        df2 = df2.rename(
            index={
                "area[1]": "area",
                "height[1]": "height",
                "sn[1]": "sn",
                "std[1]": "std",
                "mean[1]": "mean",
            }
        )
        df2["acquisition"] = len(parameters) * [f"{ui.acquisition}"]
        df2["experiment_or_precursor_mz"] = len(parameters) * [ui.precursor]
        df2["product_mz_start"] = len(parameters) * [ui.product_mz_start]
        df2["product_mz_end"] = len(parameters) * [ui.product_mz_end]
        df2["is_peak"] = is_peak
        df2["cause_for_rejection"] = rejection_cause
        df2["model_type"] = len(parameters) * [model]
        df2["subpeak"] = len(parameters) * ["2nd"]
        df_double = pandas.concat([df, df2])
        df_summary = pandas.concat([df_summary, df_double])

    else:
        # for single peak
        parameters = [
            "baseline_intercept",
            "baseline_slope",
            "mean",
            "noise",
            "std",
            "area",
            "height",
            "sn",
        ]
        df = az_summary.loc[parameters, :]
        df["acquisition"] = len(parameters) * [f"{ui.acquisition}"]
        df["experiment_or_precursor_mz"] = len(parameters) * [ui.precursor]
        df["product_mz_start"] = len(parameters) * [ui.product_mz_start]
        df["product_mz_end"] = len(parameters) * [ui.product_mz_end]
        df["is_peak"] = is_peak
        df["cause_for_rejection"] = rejection_cause
        df["model_type"] = len(parameters) * [model]
        df["subpeak"] = len(parameters) * [""]
        df_summary = pandas.concat([df_summary, df])
    # pandas.concat(df_summary, df)
    # save summary df as Excel file
    with pandas.ExcelWriter(
        path=rf"{ui.path}/peak_data_summary.xlsx", engine="openpyxl", mode="w"
    ) as writer:
        df_summary.to_excel(writer)
    return df_summary


def report_area_sheet(path: Union[str, os.PathLike], df_summary: pandas.DataFrame):
    """
    Save a different, more minimalist report sheet focussing on the area data.

    Parameters
    ----------
    path
        Path to the directory containing the raw data.
    df_summary
        DataFrame for collecting the results (i.e. peak parameters) of every signal of a given pipeline.
    """
    # also save a version of df_summary only for areas with correct order and only necessary data
    df_area_summary = df_summary[df_summary.index == "area"]
    sorted_area_summary = df_area_summary.sort_values(
        ["acquisition", "experiment_or_precursor_mz", "product_mz_start"]
    )
    sorted_area_summary = sorted_area_summary.drop(
        labels=["mcse_mean", "mcse_sd", "ess_bulk", "ess_tail"], axis=1
    )
    sorted_area_summary.to_excel(rf"{path}/area_summary.xlsx")
    return


def report_add_nan_to_summary(
    filename: str, ui: UserInput, df_summary: pandas.DataFrame, rejection_cause: str
):
    """
    Method to add NaN values to the summary DataFrame in case a signal did not contain a peak.

    Parameters
    ----------
    ui
        Instance of the UserInput class.
    df_summary
        DataFrame for collecting the results (i.e. peak parameters) of every signal of a given pipeline.
    rejection_cause
        Cause for rejecting a given signal.

    Returns
    -------
    df_summary
        Updated DataFrame for collecting the results (i.e. peak parameters) of every signal of a given pipeline.
    """
    model = ui.user_info[filename][0]
    # create DataFrame with correct format and fill it with NaN
    nan_dictionary = {
        "mean": np.nan,
        "sd": np.nan,
        "hdi_3%": np.nan,
        "hdi_97%": np.nan,
        "mcse_mean": np.nan,
        "mcse_sd": np.nan,
        "ess_bulk": np.nan,
        "ess_tail": np.nan,
        "r_hat": np.nan,
    }
    df = pandas.DataFrame(
        {
            "baseline_intercept": nan_dictionary,
            "baseline_slope": nan_dictionary,
            "mean": nan_dictionary,
            "noise": nan_dictionary,
            "std": nan_dictionary,
            "area": nan_dictionary,
            "height": nan_dictionary,
            "sn": nan_dictionary,
        }
    ).transpose()
    # add information about the signal
    df["acquisition"] = len(df.index) * [f"{ui.acquisition}"]
    df["experiment_or_precursor_mz"] = len(df.index) * [ui.precursor]
    df["product_mz_start"] = len(df.index) * [ui.product_mz_start]
    df["product_mz_end"] = len(df.index) * [ui.product_mz_end]
    df["is_peak"] = len(df.index) * [False]
    df["cause_for_rejection"] = len(df.index) * [rejection_cause]
    # if no peak was detected, there is no need for splitting double peaks, just give the info whether one was expected or not
    df["model_type"] = len(df.index) * [model]
    df["subpeak"] = len(df.index) * [""]
    # concatenate to existing summary DataFrame
    df_summary = pandas.concat([df_summary, df])
    # save summary df as Excel file
    with pandas.ExcelWriter(
        path=rf"{ui.path}/peak_data_summary.xlsx", engine="openpyxl", mode="w"
    ) as writer:
        df_summary.to_excel(writer)
    return df_summary


def pipeline_read_template(path_raw_data: Union[str, os.PathLike]):
    """
    Function to read and check the input settings and data from Template.xlsx when running the data pipeline.

    Parameters
    ----------
    path_raw_data
        Path to the raw data files. Files should be in the given raw_data_file_format, default is '.npy'.
        The `.npy` files are expected to be (2, ?)-shaped 2D NumPy arrays with time and intensity in the first dimension.

    Returns
    -------
    pre_filtering
        If True, potential peaks will be filtered based on retention time and signal to noise ratio before sampling.
    plotting
        If True, PeakPerformance will plot results.
    peak_width_estimate
        Rough estimate of the average peak width in minutes expected for the LC-MS method with which the data was obtained.
    minimum_sn
        Minimum signal to noise ratio for a signal to be recognized as a peak during pre-filtering.
    df_signals
        Read-out of the signals tab from Template.xlsx as a DataFrame.
    unique_identifiers
        List of unique identifiers from the signals tab of Template.xlsx.
    """
    # read data and user input from the settings tab of Template.xlsx
    df_settings = pandas.read_excel(
        Path(path_raw_data) / "Template.xlsx", sheet_name="settings", index_col="parameter"
    )
    pre_filtering = eval(df_settings.loc["pre_filtering", "setting"])
    if not isinstance(pre_filtering, bool):
        raise InputError("pre_filtering under settings in Template.xlsx must be a bool.")
    plotting = eval(df_settings.loc["plotting", "setting"])
    if not isinstance(plotting, bool):
        raise InputError("plotting under settings in Template.xlsx must be a bool.")
    peak_width_estimate = df_settings.loc["peak_width_estimate", "setting"]
    if not isinstance(peak_width_estimate, float) and not isinstance(peak_width_estimate, int):
        try:
            peak_width_estimate = float(peak_width_estimate)
        except:  # noqa: E722
            raise InputError(
                "peak_width_estimate under settings in Template.xlsx must be an int or float."
            )
    minimum_sn = df_settings.loc["minimum_sn", "setting"]
    if not isinstance(minimum_sn, float) and not isinstance(minimum_sn, int):
        try:
            minimum_sn = float(minimum_sn)
        except:  # noqa: E722
            raise InputError("minimum_sn under settings in Template.xlsx must be an int or float.")

    # read data and user input from the signals tab of Template.xlsx
    df_signals = pandas.read_excel(Path(path_raw_data) / "Template.xlsx", sheet_name="signals")
    unique_identifiers = list(df_signals["unique_identifier"].replace("", np.nan).dropna())
    unique_identifiers = [str(identifier) for identifier in unique_identifiers]
    if not unique_identifiers:
        raise InputError(
            "The list in column unique_identifier in the signals tab of Template.xlsx must not be empty."
        )
    if len(set(unique_identifiers)) != len(unique_identifiers):
        raise InputError(
            "The list in column unique_identifier in the signals tab of Template.xlsx must contain only unique entries."
        )
    # test whether df_signals is filled out correctly
    for x in range(len(df_signals)):
        if not df_signals.isnull()["unique_identifier"][x] and df_signals.isnull()["model_type"][x]:
            raise InputError(
                f"In the signals tab of Template.xlsx, the unique identifier in row {x + 2} has no model type."
            )
        if pre_filtering:
            if (
                not df_signals.isnull()["unique_identifier"][x]
                and df_signals.isnull()["retention_time_estimate"][x]
            ):
                raise InputError(
                    f"In the signals tab of Template.xlsx, the unique_identifier in row {x + 2} has no retention time estimate."
                )
    df_signals.set_index("unique_identifier", inplace=True)
    return pre_filtering, plotting, peak_width_estimate, minimum_sn, df_signals, unique_identifiers


def pipeline_loop(
    path_raw_data: Union[str, os.PathLike],
    path_results: Union[str, os.PathLike],
    raw_data_file_format: str,
    df_summary: pandas.DataFrame,
    *,
    restart: bool = False,
):
    """
    Function to run the complete PeakPerformance pipeline.

    Parameters
    ----------
    path_raw_data
        Path to the raw data files. Files should be in the given raw_data_file_format, default is '.npy'.
        The `.npy` files are expected to be (2, ?)-shaped 2D NumPy arrays with time and intensity in the first dimension.
    path_results
        Path to the directory for the results of a given Batch run of PeakPerformance.
    raw_data_file_format
        Data format (suffix) of the raw data, default is '.npy'.
    df_summary
        DataFrame for collecting the results (i.e. peak parameters) of every signal of a given pipeline.
    restart
        If a pipeline broke for some reason, it can be restarted by setting restart to True.
        That way, already analyzed files won't be analyzed again.
    """
    # read data and user input from the settings tab of Template.xlsx
    (
        pre_filtering,
        plotting,
        peak_width_estimate,
        minimum_sn,
        df_signals,
        unique_identifiers,
    ) = pipeline_read_template(path_raw_data)
    peak_model_list = []
    retention_time_estimate_list = []
    # synchronize the lists of raw data files, peak models, and retention times
    # they will be converted to the user_info dict when instantiating the UserInput class below
    df_files = pandas.read_excel(Path(path_raw_data) / "Template.xlsx", sheet_name="files")
    raw_data_files = list(df_files.loc[:, "file_name"])
    # in case of a restart, update raw_data_files to only contain files which have not been analyzed
    if restart:
        analyzed_files = os.listdir(path_results)
        for raw in raw_data_files:
            for analyzed in analyzed_files:
                if raw in analyzed:
                    raw_data_files.remove(raw)
    for file in raw_data_files:
        for identifier in unique_identifiers:
            if identifier in file:
                peak_model_list.append(str(df_signals.loc[identifier, "model_type"]))
                retention_time_estimate_list.append(
                    df_signals.loc[identifier, "retention_time_estimate"]
                )

    # loop over filenames
    for file in raw_data_files:
        # parse the data and extract information from the (standardized) file name
        (
            timeseries,
            acquisition,
            precursor,
            product_mz_start,
            product_mz_end,
        ) = parse_data(path_raw_data, file, raw_data_file_format)
        # instantiate the UserInput class all given information
        ui = UserInput(
            path_results,
            raw_data_files,
            raw_data_file_format,
            peak_model_list,
            retention_time_estimate_list,
            peak_width_estimate,
            pre_filtering,
            minimum_sn,
            timeseries,
            acquisition,
            precursor,
            product_mz_start,
            product_mz_end,
        )
        # apply pre-sampling filter (if selected)
        if pre_filtering:
            # test if necessary settings were provided by the user
            if not retention_time_estimate_list:
                raise InputError(
                    "If selecting pre-filtering, provide a list of retention time estimate in Template.xlsx."
                )
            if not minimum_sn:
                raise InputError(
                    "If selecting pre-filtering, provide a minimum signal-to-noise ratio in Template.xlsx."
                )
            if not peak_width_estimate:
                raise InputError(
                    "If selecting pre-filtering, provide a rough estimate of the general peak width in Template.xlsx."
                )

            # calculate noise guess for pre-filtering
            slope_guess, intercept_guess, noise_guess = models.initial_guesses(
                ui.timeseries[0], ui.timeseries[1]
            )
            prefilter, df_summary = prefiltering(file, ui, noise_guess, df_summary)
            if not prefilter:
                # if no peak candidates were found, continue with the next signal
                if plotting:
                    plots.plot_raw_data(
                        file[: -len(ui.raw_data_file_format)],
                        ui.timeseries[0],
                        ui.timeseries[1],
                        ui.path,
                    )
                continue
        # select model based on information in UserInput
        model = ui.user_info[file][0]
        if model == models.ModelType.Normal:
            pmodel = models.define_model_normal(ui.timeseries[0], ui.timeseries[1])
        elif model == models.ModelType.SkewNormal:
            pmodel = models.define_model_skew(ui.timeseries[0], ui.timeseries[1])
        elif model == models.ModelType.DoubleNormal:
            pmodel = models.define_model_double_normal(ui.timeseries[0], ui.timeseries[1])
        elif model == models.ModelType.DoubleSkewNormal:
            pmodel = models.define_model_double_skew_normal(ui.timeseries[0], ui.timeseries[1])
        else:
            raise NotImplementedError(
                f"The model '{model}' specified for file '{file}' is not implemented."
            )

        # sample the chosen model
        idata = sampling(pmodel)
        # apply post-sampling filter
        resample, discard, df_summary = postfiltering(file, idata, ui, df_summary)
        # if peak was discarded, continue with the next signal
        if discard:
            if plotting:
                plots.plot_posterior(
                    file[: -len(ui.raw_data_file_format)],
                    ui.timeseries[0],
                    ui.timeseries[1],
                    ui.path,
                    idata,
                    True,
                )
            continue
        # if convergence was not yet reached, sample again with more tuning samples
        if resample:
            if "double" in model:
                idata = sampling(pmodel, tune=16000)
            else:
                idata = sampling(pmodel, tune=6000)
            resample, discard, df_summary = postfiltering(file, idata, ui, df_summary)
            if discard:
                plots.plot_posterior(
                    file[: -len(ui.raw_data_file_format)],
                    ui.timeseries[0],
                    ui.timeseries[1],
                    ui.path,
                    idata,
                    True,
                )
                continue
            if resample:
                # if signal was flagged for re-sampling a second time, discard it
                rejection_msg = "postfiltering: signal was flagged for re-sampling with increased sample number twice"
                df_summary = report_add_data_to_summary(
                    file, idata, df_summary, ui, False, rejection_msg
                )
                if plotting:
                    plots.plot_posterior(
                        file[: -len(ui.raw_data_file_format)],
                        ui.timeseries[0],
                        ui.timeseries[1],
                        ui.path,
                        idata,
                        True,
                    )
                continue
        # perform posterior predictive sampling
        idata = posterior_predictive_sampling(pmodel, idata)
        # add inference data to df_summary and save it as an Excel file
        df_summary = report_add_data_to_summary(file, idata, df_summary, ui, True)
        # save the inference data object as a netcdf file
        report_save_idata(idata, ui, file[: -len(ui.raw_data_file_format)])
        # plot data
        if plotting:
            plots.plot_posterior_predictive(
                file[: -len(ui.raw_data_file_format)],
                ui.timeseries[0],
                ui.timeseries[1],
                ui.path,
                idata,
                False,
            )
            plots.plot_posterior(
                file[: -len(ui.raw_data_file_format)],
                ui.timeseries[0],
                ui.timeseries[1],
                ui.path,
                idata,
                False,
            )
        # save condesed Excel file with area data
        report_area_sheet(path_results, df_summary)


def pipeline(
    path_raw_data: Union[str, os.PathLike],
    raw_data_file_format: str,
):
    """
    Function to run the complete PeakPerformance pipeline.

    Parameters
    ----------
    path_raw_data
        Path to the raw data files. Files should be in the given raw_data_file_format, default is '.npy'.
        The `.npy` files are expected to be (2, ?)-shaped 2D NumPy arrays with time and intensity in the first dimension.
    raw_data_file_format
        Data format (suffix) of the raw data, default is '.npy'.

    Returns
    ----------
    path_results
        Path variable pointing to the newly created folder for this batch.
    """
    # create data structure and DataFrame(s) for results
    df_summary, path_results = initiate(path_raw_data)
    pipeline_loop(
        path_raw_data,
        path_results,
        raw_data_file_format,
        df_summary,
    )
    return path_results


def pipeline_restart(
    path_raw_data: Union[str, os.PathLike],
    raw_data_file_format: str,
    path_results: Union[str, os.PathLike],
):
    """
    Function to restart a broken PeakPerformance pipeline.
    Files which are in the results directory of the broken pipeline will not be analyzed again.
    WARNING: This only works once! If a pipeline fails more than once, copy all files (except the Excel report sheets)
    into one directory and specify this directory as the path_results argument.

    Parameters
    ----------
    path_raw_data
        Path to the raw data files. Files should be in the given raw_data_file_format, default is '.npy'.
        The `.npy` files are expected to be (2, ?)-shaped 2D NumPy arrays with time and intensity in the first dimension.
    raw_data_file_format
        Data format (suffix) of the raw data, default is '.npy'.
    path_results
        Path variable pointing to the directory of the broken PeakPerformance batch

    Returns
    ----------
    path_results_new
        Path variable pointing to the newly created folder for the restarted batch.
    """
    df_summary, path_results_new = initiate(path_raw_data)
    pipeline_loop(
        path_raw_data,
        path_results,
        raw_data_file_format,
        df_summary,
        restart=True,
    )
    return path_results_new


def excel_template_prepare(
    path_raw_data: Union[str, os.PathLike],
    path_peak_performance: Union[str, os.PathLike],
    raw_data_files: Union[List[str], Tuple[str]],
    unique_identifiers: Union[List[str], Tuple[str]],
):
    """
    Function to copy Template.xlsx from the peak performance directory to the directory containing the raw data files.
    Subsequently, update Template.xlsx with a list of all raw data files and of all unique_identifiers.

    Parameters
    ----------
    path_raw_data
        Path to the folder containing raw data.
    path_peak_performance
        Path to the folder containing PeakPerformance.
    raw_data_files
        List with names of all files of the specified data type in path_raw_data.
    unique_identifiers
        List with all unique combinations of targeted molecules.
        (i.e. experiment number or precursor ion m/z ratio and product ion m/z ratio range)
    """
    # copy Template.xlsx from PeakPerformance to the directory with the raw data
    try:
        shutil.copy(
            Path(path_peak_performance) / "Template.xlsx", Path(path_raw_data) / "Template.xlsx"
        )
    except FileNotFoundError:
        raise ParsingError(f"Template.xlsx was not found in {path_peak_performance}.")
    except Exception:
        raise ParsingError(
            f"Error while copying Template.xlsx from {path_peak_performance} into {path_raw_data}."
        )
    # load Template.xlsx
    wb = load_workbook(Path(path_raw_data) / "Template.xlsx")
    # add list of all files names to the files tab
    wb_files = wb["files"]
    df1 = pandas.DataFrame({"file_name": raw_data_files})
    for r in dataframe_to_rows(df1, index=False, header=False):
        wb_files.append(r)
    # add list of all unique identifiers (i.e. mass traces) to the signals tab
    wb_signals = wb["signals"]
    df2 = pandas.DataFrame({"unique_identifier": unique_identifiers})
    for r in dataframe_to_rows(df2, index=False, header=False):
        wb_signals.append(r)
    wb.save(Path(path_raw_data) / "Template.xlsx")
    return


def prepare_model_selection(
    path_raw_data: Union[str, os.PathLike],
    path_template: Union[str, os.PathLike],
):
    """
    Function to prepare model selection by providing and mostly filling out an Excel template
    Template.xlsx. After this step, the user has to provide relevant information in Template.xlsx
    which is finally used for model selection.

    Parameters
    ----------
    path_raw_data
        Path to the folder containing raw data.
    path_template
        Path to the folder containing Template.xlsx from PeakPerformance.
    """
    # detect raw data files
    raw_data_files = detect_raw_data(path_raw_data)
    # parse unique identifiers
    identifiers = parse_unique_identifiers(raw_data_files)
    # copy Template.xlsx into raw data directory and add data from the previous commmands
    excel_template_prepare(path_raw_data, path_template, raw_data_files, identifiers)
    return


def parse_files_for_model_selection(signals: pandas.DataFrame) -> Dict[str, str]:
    """
    Function to parse the file names for model selection.

    Parameters
    ----------
    signals
        DataFrame containing the signals tab of Template.xlsx.

    Returns
    ----------
    files_for_selection
        Dict with file names as keys and unique identifiers as values.
    """
    identifier_list = list(signals["unique_identifier"].replace("", np.nan).dropna())
    model_list = list(signals["model_type"].replace("", np.nan).dropna())
    acquisition_list = list(
        signals["acquisition_for_choosing_model_type"].replace("", np.nan).dropna()
    )
    # sanity checks
    if not identifier_list:
        raise InputError("In the signals tab of Template.xlsx, there are no unqiue_identifiers.")
    if not model_list and not acquisition_list:
        raise InputError(
            "In the signals tab of Template.xlsx, no model or acquisition(s) for model selection were provided."
        )
    if len(identifier_list) == len(model_list):
        raise InputError(
            """In the signals tab of Template.xlsx, for each unique identifier a model type was provided.
Thus, no model selection is performed."""
        )
    # multiple scenarios have to be covered
    files_for_selection: Dict[str, str] = {}
    signals = signals.fillna("")
    if len(model_list) == len(signals.index):
        # scenario 1: a model was specified for every unique identifier (by the user) -> model selection obsolete
        return files_for_selection
    elif len(signals.index) - len(model_list) > 1 and len(acquisition_list) == 1:
        # scenario 2: for more than one unique identifier no model was specified by the user
        # but a single acquisition was given for model selection -> model selection from one acquisition
        acquisition = acquisition_list[0]
        # remove possible whitespace in front or after an entry made by the user
        acquisition = acquisition.strip()
        for idx, row in signals.iterrows():
            if not signals.loc[idx, "model_type"]:
                unique_identifier = getattr(row, "unique_identifier")
                filename = "_".join([acquisition, unique_identifier])
                files_for_selection[filename] = unique_identifier
    elif len(signals.index) - len(model_list) == len(acquisition_list):
        # scenario 3: for every unique identifier for which no model was specified by the user,
        # they provided an acquistion for model selection
        for idx, row in signals.iterrows():
            if not signals.loc[idx, "model_type"]:
                acquisition = getattr(row, "acquisition_for_choosing_model_type")
                unique_identifier = getattr(row, "unique_identifier")
                filename = "_".join([acquisition, unique_identifier])
                files_for_selection[filename] = unique_identifier
    else:
        raise InputError(
            "When using model selection, provide either one acquisition or one acquisition per unique identifier (no in-betweens)."
        )
    return files_for_selection


def selected_models_to_template(
    path_raw_data: Union[str, os.PathLike],
    signals: pandas.DataFrame,
    model_dict: Mapping[str, str],
):
    """
    Function to update Template.xlsx with the selected model types.

    Parameters
    ----------
    path_raw_data
        Path to the folder containing raw data.
    signals
        DataFrame containing the signals tab of Template.xlsx.
    model_dict
        Dict with unique identifiers as keys and model types as values.
    """
    signals = signals.fillna("")
    for idx, row in signals.iterrows():
        if not signals.loc[idx, "model_type"]:
            unique_identifier = getattr(row, "unique_identifier")
            signals.loc[idx, "model_type"] = model_dict[unique_identifier]
    # update in Excel
    wb = load_workbook(Path(path_raw_data) / "Template.xlsx")
    # update signals tab with model types by deleting rows and appending signals
    wb_signals = wb["signals"]
    wb_signals.delete_rows(wb_signals.min_row + 1, wb_signals.max_row)
    for r in dataframe_to_rows(signals, index=False, header=False):
        wb_signals.append(r)
    wb.save(Path(path_raw_data) / "Template.xlsx")
    return


def model_selection_check(
    result_df: pandas.DataFrame, ic: str, elpd_threshold: Union[str, float] = 25
) -> str:
    """
    During model seleciton, double peak models are sometimes incorrectly preferred due to their increased complexity.
    Therefore, they have to outperform single peak models by an empirically determined value of the elpd.

    Parameters
    ----------
    result_df
        DataFrame with the result of model comparison via az.compare().
    ic
        Information criterion to be used for model selection.
        ("loo": pareto-smoothed importance sampling leave-one-out cross-validation,
        "waic": widely applicable information criterion)
    elpd_threshold
        Threshold of the elpd difference between a double and a single peak model for the double peak model
        to be accepted.

    Returns
    ----------
    selected_model
        Name of the selected model type.
    """
    selected_model = str(result_df.index[0])
    if "double" in selected_model:
        df_single_peak_models = result_df[~result_df.index.str.contains("double")]
        elpd_single = max(list(df_single_peak_models[f"elpd_{ic}"]))
        elpd_double = max(list(result_df[f"elpd_{ic}"]))
        if not elpd_double > elpd_single + elpd_threshold:
            selected_model = str(df_single_peak_models.index[0])
    return selected_model


def selection_loop(
    path_raw_data: Union[str, os.PathLike],
    *,
    files_for_selection: Mapping[str, str],
    raw_data_files: Union[List[str], Tuple[str]],
    ic: str,
    signals: pandas.DataFrame,
):
    """
    Function containing the loop over all filenames intended for the model selection.
    Involves sampling every model featured by PeakPerformance, computing the loglikelihood
    and an information criterion, and comparing the results to ascertain the best model for every file.

    Parameters
    ----------
    path_raw_data
        Path to the folder containing raw data.
    files_for_selection
        Dict with file names as keys and unique identifiers as values.
    raw_data_files
        List of raw data files returned by the detect_raw_data() function.
        Is needed here only to get access to the file format.
    ic
        Information criterion to be used for model selection.
        ("loo": pareto-smoothed importance sampling leave-one-out cross-validation,
        "waic": widely applicable information criterion)

    Returns
    ----------
    result_df
        DataFrame containing the ranking and scores of the model selection.
    model_dict
        Dict with unique identifiers as keys and model types as values.
    """
    model_dict = {}
    # get data file format from raw_data_files
    file_format = raw_data_files[0].split(".")[-1]
    # loop over all filenames in files_for_selection
    for filename in files_for_selection.keys():
        # load time series
        timeseries = np.load(Path(path_raw_data) / (filename + "." + file_format))
        idata_dict = {}
        # get all implemented models, then remove those which were excluded
        # from model selection by the user
        models_to_exclude = str(
            signals.loc[files_for_selection[filename], "models_to_exclude_from_selection"]
        )
        model_list = set(models.ModelType)
        if models_to_exclude:
            exclude_models = {mex.strip() for mex in models_to_exclude.split(",")}
            model_list = model_list - exclude_models  # type: ignore[operator]
        if models.ModelType.Normal in model_list:
            pmodel_normal = models.define_model_normal(timeseries[0], timeseries[1])
            idata_normal = sampling(pmodel_normal, tune=6000)
            idata_normal = models.compute_log_likelihood(pmodel_normal, idata_normal)
            idata_normal_summary = az.summary(idata_normal)
            idata_dict["normal"] = [idata_normal_summary, idata_normal]
        if models.ModelType.SkewNormal in model_list:
            pmodel_skew = models.define_model_skew(timeseries[0], timeseries[1])
            idata_skew = sampling(pmodel_skew, tune=6000)
            idata_skew = models.compute_log_likelihood(pmodel_skew, idata_skew)
            idata_skew_normal_summary = az.summary(idata_skew)
            idata_dict["skew_normal"] = [idata_skew_normal_summary, idata_skew]
        if models.ModelType.DoubleNormal in model_list:
            pmodel_double_normal = models.define_model_double_normal(timeseries[0], timeseries[1])
            idata_double_normal = sampling(pmodel_double_normal, tune=6000)
            idata_double_normal = models.compute_log_likelihood(
                pmodel_double_normal, idata_double_normal
            )
            idata_double_normal_summary = az.summary(idata_double_normal)
            idata_dict["double_normal"] = [idata_double_normal_summary, idata_double_normal]
        if models.ModelType.DoubleSkewNormal in model_list:
            pmodel_double_skew = models.define_model_double_skew_normal(
                timeseries[0], timeseries[1]
            )
            idata_double_skew = sampling(pmodel_double_skew, tune=6000)
            idata_double_skew = models.compute_log_likelihood(pmodel_double_skew, idata_double_skew)
            idata_double_skew_normal_summary = az.summary(idata_double_skew)
            idata_dict["double_skew_normal"] = [idata_double_skew_normal_summary, idata_double_skew]

        # add model to compare_dict for model selection only if convergence criterion was met (r_hat <= 1.05)
        compare_dict = {}
        for model in idata_dict.keys():
            if not (idata_dict[model][0].loc[:, "r_hat"] > 1.05).any():
                compare_dict[model] = idata_dict[model][1]
        # compare_dict needs at least two entries for model comparison
        # if not enough pass the r_hat test, accept all for now to avoid error
        if len(compare_dict) < 2:
            warnings.warn(
                f"Only one or less models converged during model selection for {filename}."
            )
            for model in idata_dict.keys():
                compare_dict[model] = idata_dict[model][1]
        # perform the actual model comparison
        result_df = models.model_comparison(compare_dict, ic)
        # double peak models are sometimes incorrectly preferred due to their increased complexity
        # therefore, they have to outperform single peak models by an empirically determined value of the elpd
        selected_model = model_selection_check(result_df, ic)
        # update model_dict with unique_identifier as key and selected_model as value
        model_dict[files_for_selection[filename]] = selected_model
        # optional: plot the results of model comparison
    return result_df, model_dict


def model_selection(path_raw_data: Union[str, os.PathLike], *, ic: str = "loo"):
    """
    Method to select the best model for every signal (i.e. combination of experiment number or precursor ion m/z ratio
    and product ion m/z ratio). This is realized by analyzing one representative sample of the batch with all models and
    comparing the results based on an informantion criterion.

    Parameters
    ----------
    path_raw_data
        Path to the folder containing raw data.
    ic
        Information criterion to be used for model selection.
        ("loo": pareto-smoothed importance sampling leave-one-out cross-validation,
        "waic": widely applicable information criterion)

    Returns
    ----------
    comparison_results
        DataFrame containing all rankings from model selection.
    model_dict
        Dict with unique identifiers as keys and model types as values.
    """
    # check for which signals model selection is wished and whether from one or different acquisitions
    df_signals = pandas.read_excel(Path(path_raw_data) / "Template.xlsx", sheet_name="signals")
    files_for_selection = parse_files_for_model_selection(df_signals)
    # get raw_data_files to get automatic access to file format in seleciton_loop
    raw_data_files = detect_raw_data(path_raw_data)
    # loop over all files_for_selection
    df_signals.set_index("unique_identifier", inplace=True)
    comparison_results = pandas.DataFrame()
    result_df, model_dict = selection_loop(
        path_raw_data,
        files_for_selection=files_for_selection,
        raw_data_files=raw_data_files,
        ic=ic,
        signals=df_signals,
    )
    comparison_results = pandas.concat([comparison_results, result_df])
    # update signals tab of Template.xlsx; read again to reset index
    df_signals = pandas.read_excel(Path(path_raw_data) / "Template.xlsx", sheet_name="signals")
    try:
        selected_models_to_template(path_raw_data, df_signals, model_dict)
    except PermissionError:
        warnings.warn(
            """Since Template.xlsx was open during model selection, it could not be updated.
Use the returned variables and pl.selected_models_to_template() to update it."""
        )
    return comparison_results, model_dict
